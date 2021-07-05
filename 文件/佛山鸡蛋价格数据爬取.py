import requests
import json
from urllib import parse
import time, datetime
import openpyxl
import random


#生成表单数据
def getData(pageNo):
    #商品信息和ID
    data="craftName=%E7%BA%A2%E7%9A%AE%E9%B8%A1%E8%9B%8B&craftIndex=13263"
    #市场名
    data+="&eudName=%E4%BD%9B%E5%B1%B1%E4%B8%AD%E5%8D%97%E5%86%9C%E4%BA%A7%E5%93%81%E4%BA%A4%E6%98%93%E4%B8%AD%E5%BF%83"
    #日期的方式和时间范围
    data+="&queryDateType=4&timeRange=2019-05-01+~+2021-01-01"
    #页数
    data+="&pageNo="+str(pageNo)
    return data

#修改时间格式
def timeStyle(timeArray): 
    timeArray = time.localtime(timeArray/1000)
    otherStyleTime = time.strftime("%Y-%m-%d", timeArray)
    return otherStyleTime

#获取服务器返回数据
def getResultSet(pageNo):
    url ="http://nc.mofcom.gov.cn/jghq/marketDetail?eudId=70801139"
    data=getData(pageNo)
    #print(data)
    headers = {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Accept-Encoding": "gzip, deflate",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,zh-TW;q=0.7",
        "Cache-Control": "no-cache",
        "Connection": "keep-alive",
        "Content-Length": "149",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "Cookie":"_pk_ref.134.0d5f=%5B%22%22%2C%22%22%2C1606394667%2C%22https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3DCz7miL7z63uLInJIhhOzky3uPTSyLQAcXgMpAS30psD7ZFUKpPHywQPnINuh67wPvI0b2Ql-4MYhgrKUk8soSeXY9XNcFrPoQ7p3YmCLq1jPJp9MnNduqywYFzPE6hH6R06THZBF3FjrlXgQAI1fNplNaHrC01xTjVDnRjLjIP3G63SLZn0sIaiwjp1nKzeGE1yyH35TAqqyGr5sPkbPOp6XmWH4R7_2z6aDjn4n5-S%26wd%3D%26eqid%3De5ad73e000057320000000065fbf44c4%22%5D; _pk_id.134.0d5f=43e24b95cd961b78.1606135259.5.1606394877.1606373978.; insert_cookie=28314071",
        "Host": "nc.mofcom.gov.cn",
        "Origin":"http://nc.mofcom.gov.cn",
        "Pragma": "no-cache",
        "Referer":"http://nc.mofcom.gov.cn/nc/qyncp/list?pIndex=440881",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36",
        "X-Requested-With": "XMLHttpRequest"
        }
    response = requests.post(url, data=data, headers=headers)
    resultSet=response.json()
    return resultSet['result']


#主程序
workbook=openpyxl.Workbook()
worksheet = workbook.active
worksheet.title="mysheet"
row=0
#鸡蛋数据共36页
for i in range(1,37):
    resultSet=getResultSet(i)
    for j in range(len(resultSet)):
        #日期，将时间戳换成当地时间格式
        date=timeStyle(resultSet[j]['GET_P_DATE'])
        #价格
        pricr=resultSet[j]['AG_PRICE']
        #单位
        unit=resultSet[j]['C_UNIT']
        row=row+1
        worksheet.cell(row,1,date)
        worksheet.cell(row,2,pricr)
        worksheet.cell(row,3,unit)
        print("第%d页，第%d行，第%d条数据，时间:%s，价格:%.2f%s" %(i,j+1,row,date,pricr,unit))
    workbook.save(filename='data.xlsx')
    sleepTime=random.randint(8,15)
    print("第%s页结果保存完毕，休息%d秒" %(i,sleepTime))
    time.sleep(sleepTime)
    
    



